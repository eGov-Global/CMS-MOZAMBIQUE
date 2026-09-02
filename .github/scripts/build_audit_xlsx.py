#!/usr/bin/env python3
"""
Build a management-ready, multi-sheet security AUDIT workbook (.xlsx) from run.json.

Sheets:
  1. Summary          - posture headline + counts by status / priority / severity / category
  2. Action Required  - the tracking list (only findings that need fixing), with blank
                        Status / Owner / Target date / Notes columns for the team
  3. Not Tracked      - acceptable ("okayish") + false positives, with the reason excluded
  4. All Locations    - every occurrence of an action-required finding (path:line + link)

Design goals: calm, professional palette (no loud full-cell fills), soft priority/severity
chips, columns auto-sized to content with padding so nothing is clipped, wrapped long text
with matching row heights, and short hyperlinked labels for references/locations.

Triage taxonomy (enrich_report.py): triage.status = action_required | acceptable |
false_positive; triage.priority = P0 | P1 | P2 | P3 (action_required only).

Env: RUN_JSON (default run.json), OUT_XLSX (default security-audit.xlsx). Requires openpyxl.
"""
import os, sys, json, math

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except Exception as e:
    print(f"openpyxl unavailable ({e}); skipping xlsx audit export.", file=sys.stderr)
    sys.exit(0)

RUN = os.environ.get("RUN_JSON", "run.json")
OUT = os.environ.get("OUT_XLSX", "security-audit.xlsx")

try:
    run = json.load(open(RUN))
except Exception as e:
    print(f"cannot read {RUN}: {e}", file=sys.stderr); sys.exit(0)

meta = run.get("meta", {})
findings = run.get("findings", [])
for _f in findings:  # fold any legacy Info into Low (no Info tier)
    if _f.get("severity") == "INFO":
        _f["severity"] = "LOW"
SEV_ORDER = ["CRITICAL", "HIGH", "MEDIUM", "LOW"]
PRI_ORDER = {"P0": 0, "P1": 1, "P2": 2, "P3": 3, "": 4, None: 4}

# ---- palette (calm, management-facing) -------------------------------------
NAVY = "1F3A5F"; INK = "1F2A37"; SUB = "6B7280"; LINE = "DDE3EC"; TITLE_BLUE = "1F3A5F"
# Priority: solid chip (strong label) + soft row background, distinct family from severity.
PRI_SOLID = {"P0": "5B21B6", "P1": "C0267E", "P2": "047857", "P3": "5F6673"}
PRI_SOFT = {"P0": "EEE8FB", "P1": "FAE7F2", "P2": "E2F3EC", "P3": "ECEEF1"}
# Severity: soft chip (fill) + solid text, darkened just enough for legibility on the tint.
SEV_CHIP = {"CRITICAL": ("F8E7EA", "7A1F2B"), "HIGH": ("FCEAEA", "B23A3A"),
            "MEDIUM": ("FFF1CC", "8A5A06"), "LOW": ("EAF0FF", "2F6FED")}
NEUTRAL_SOFT = "F6F7FA"  # row background for Not Tracked (no priority)


def status_of(f): return (f.get("triage") or {}).get("status") or "action_required"
def priority_of(f): return (f.get("triage") or {}).get("priority") or ("P2" if status_of(f) == "action_required" else "")


action = [f for f in findings if status_of(f) == "action_required"]
acceptable = [f for f in findings if status_of(f) == "acceptable"]
false_pos = [f for f in findings if status_of(f) == "false_positive"]
action.sort(key=lambda f: (PRI_ORDER.get(priority_of(f), 4),
                           SEV_ORDER.index(f["severity"]) if f["severity"] in SEV_ORDER else 9, -f.get("count", 0)))

# ---- style helpers ---------------------------------------------------------
THIN = Side(style="thin", color=LINE)
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEAD_FILL = PatternFill("solid", fgColor=NAVY)
HEAD_FONT = Font(color="FFFFFF", bold=True, size=10.5, name="Calibri")
WRAP = Alignment(vertical="top", wrap_text=True)
TOP = Alignment(vertical="top")
CTR = Alignment(horizontal="center", vertical="center")


def chip(cell, text, fill, font):
    cell.value = text
    cell.fill = PatternFill("solid", fgColor=fill)
    cell.font = Font(color=font, bold=True, size=10)
    cell.alignment = CTR
    cell.border = BORDER


def solid_chip(cell, text, fill):
    cell.value = text
    cell.fill = PatternFill("solid", fgColor=fill)
    cell.font = Font(color="FFFFFF", bold=True, size=10)
    cell.alignment = CTR
    cell.border = BORDER


def header(ws, cols, row=1):
    for c, spec in enumerate(cols, 1):
        cell = ws.cell(row=row, column=c, value=spec["h"])
        cell.fill = HEAD_FILL; cell.font = HEAD_FONT
        cell.alignment = Alignment(horizontal=spec.get("align", "left"), vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[row].height = 30
    ws.freeze_panes = ws.cell(row=row + 1, column=1)


def autosize(ws, cols, data_rows):
    """Set column widths from content length (+padding), capped per spec, and compute
    row heights so wrapped cells show fully."""
    for ci, spec in enumerate(cols, 1):
        w = spec.get("w")
        if w is None:  # auto from content
            maxlen = len(str(spec["h"]))
            for r in range(2, data_rows + 2):
                v = ws.cell(row=r, column=ci).value
                if v is not None:
                    maxlen = max(maxlen, max((len(x) for x in str(v).split("\n")), default=0))
            w = min(spec.get("max", 40), max(spec.get("min", 9), maxlen + 3))
        ws.column_dimensions[get_column_letter(ci)].width = w
    # row heights for wrapped columns
    wrapcols = [(ci, spec) for ci, spec in enumerate(cols, 1) if spec.get("wrap")]
    for r in range(2, data_rows + 2):
        lines = 1
        for ci, spec in enumerate(cols, 1):
            if not spec.get("wrap"):
                continue
            v = ws.cell(row=r, column=ci).value
            if v is None:
                continue
            width = ws.column_dimensions[get_column_letter(ci)].width or 20
            chars = max(8, width / 1.05)
            need = sum(max(1, math.ceil(len(seg) / chars)) for seg in str(v).split("\n"))
            lines = max(lines, need)
        ws.row_dimensions[r].height = min(300, 15 * lines + 6)


def link(cell, text, url):
    cell.value = text
    if url:
        cell.hyperlink = url
        cell.font = Font(color="1D4ED8", underline="single", size=10)
    cell.alignment = TOP


wb = Workbook()

# ============================ Sheet 1: Summary ==============================
ws = wb.active; ws.title = "Summary"
ws.sheet_view.showGridLines = False
ws["A1"] = "Security Audit — Ansible Remote Server Deployment"
ws["A1"].font = Font(color=TITLE_BLUE, bold=True, size=16)
ws.merge_cells("A1:C1")
ws["A2"] = f"{meta.get('repo','')}   ·   branch {meta.get('branch','')}   ·   commit {meta.get('shaShort','')}   ·   scanned {meta.get('date','')}"
ws["A2"].font = Font(color=SUB, size=10); ws.merge_cells("A2:C2")
eng = meta.get("engine")
ws["A3"] = (f"Triage & remediation: {eng}, fail-safe dual-pass verified." if eng else "Remediation: curated ruleset (AI enrichment not run).")
ws["A3"].font = Font(color=SUB, size=10, italic=True); ws.merge_cells("A3:C3")

# posture headline
p0 = sum(1 for f in action if priority_of(f) == "P0")
high = sum(1 for f in action if f["severity"] in ("CRITICAL", "HIGH"))
posture = "ACTION REQUIRED" if action else "NO ISSUES"
pfill = "C0392B" if p0 or high else ("B7791F" if action else "1E8449")
ws["A5"] = posture
ws["A5"].font = Font(color="FFFFFF", bold=True, size=12); ws["A5"].alignment = CTR
ws["A5"].fill = PatternFill("solid", fgColor=pfill)
ws.merge_cells("A5:C5"); ws.row_dimensions[5].height = 24


def block(ws, title, rows, r):
    ws.cell(row=r, column=1, value=title).font = Font(bold=True, color=NAVY, size=11)
    r += 1
    for label, val in rows:
        a = ws.cell(row=r, column=1, value=label); a.font = Font(size=10.5); a.alignment = TOP; a.border = BORDER
        b = ws.cell(row=r, column=2, value=val); b.font = Font(bold=True, size=10.5); b.alignment = Alignment(horizontal="right"); b.border = BORDER
        r += 1
    return r + 1


r = 7
r = block(ws, "Tracking status", [
    ("Action required (tracked)", len(action)),
    ("Acceptable / not tracked", len(acceptable)),
    ("False positives", len(false_pos)),
    ("Total issue types", len(findings)),
    ("Total occurrences", run.get("summary", {}).get("occurrences", sum(f.get("count", 0) for f in findings))),
], r)
r = block(ws, "Action required by priority", [(p, sum(1 for f in action if priority_of(f) == p)) for p in ("P0", "P1", "P2", "P3")], r)
cats = {}
for f in action:
    cats[f.get("category", "General Hardening")] = cats.get(f.get("category", "General Hardening"), 0) + 1
r = block(ws, "Action required by category", sorted(cats.items(), key=lambda x: -x[1]), r)
ws.column_dimensions["A"].width = 34; ws.column_dimensions["B"].width = 12; ws.column_dimensions["C"].width = 4

# ======================= Sheet 2: Action Required ===========================
ws = wb.create_sheet("Action Required")
ws.sheet_view.showGridLines = False
COLS = [
    {"h": "Priority", "align": "center", "w": 9},
    {"h": "Severity", "align": "center", "w": 11},
    {"h": "Category", "wrap": True, "min": 16, "max": 22, "w": 20},
    {"h": "Finding", "wrap": True, "w": 34},
    {"h": "Rule ID", "min": 12, "max": 40},
    {"h": "Scanner", "w": 12},
    {"h": "Exposure", "align": "center", "w": 11},
    {"h": "Count", "align": "center", "w": 8},
    {"h": "Why it matters", "wrap": True, "w": 52},
    {"h": "How to fix", "wrap": True, "w": 54},
    {"h": "Reference", "w": 13},
    {"h": "Location", "min": 20, "max": 46},
    {"h": "Status", "w": 13},
    {"h": "Owner", "w": 14},
    {"h": "Target date", "w": 13},
    {"h": "Notes", "wrap": True, "w": 30},
]
header(ws, COLS)
row = 2
for f in action:
    tri = f.get("triage") or {}
    loc = (f.get("locations") or [{}])[0]
    pr = priority_of(f)
    rf = PatternFill("solid", fgColor=PRI_SOFT.get(pr, NEUTRAL_SOFT))  # whole-row tint = priority soft
    why = f.get("why", "")
    sx = f.get("strix") or ({"cvss": f.get("cvss"), "cwe": f.get("cwe")} if f.get("source") == "Strix" else None)
    if sx:  # surface independent Strix AI-pentest validation + CVSS/CWE inline
        tag = "[Strix-validated" + (f" · CVSS {sx.get('cvss')}" if sx.get("cvss") else "") + (f" · {sx.get('cwe')}" if sx.get("cwe") else "") + "]  "
        why = tag + why
    vv = f.get("validation")
    if vv and vv.get("adjusted") and vv.get("scanner_severity"):  # LLM severity calibration audit trail
        why = f"[AI-validated · assessed {str(vv.get('severity','')).title()} (scanner rated {str(vv['scanner_severity']).title()})]  " + why
    vals = {3: f.get("category", ""), 4: f.get("title", ""), 5: f.get("id", ""), 6: f.get("source", ""),
            7: tri.get("exposure", ""), 8: f.get("count", 0), 9: why, 10: f.get("fix", ""),
            13: "Open", 16: tri.get("reason", "")}
    for c, v in vals.items():
        cell = ws.cell(row=row, column=c, value=v)
        cell.alignment = WRAP if COLS[c - 1].get("wrap") else TOP
        cell.border = BORDER; cell.fill = rf
    lk = ws.cell(row=row, column=11); link(lk, "Reference ↗" if f.get("guide") else "", f.get("guide")); lk.border = BORDER; lk.fill = rf
    lc = ws.cell(row=row, column=12); link(lc, f"{loc.get('path','')}:{loc.get('line','')}" if loc.get("path") else "", loc.get("url")); lc.border = BORDER; lc.fill = rf
    for c in (14, 15):
        cc = ws.cell(row=row, column=c); cc.border = BORDER; cc.fill = rf
    solid_chip(ws.cell(row=row, column=1), pr, PRI_SOLID.get(pr, "5F6673"))
    chip(ws.cell(row=row, column=2), f["severity"].title(), *SEV_CHIP.get(f["severity"], ("EAEDF2", "475569")))
    row += 1
if row == 2:
    ws.cell(row=2, column=1, value="No action-required findings.").font = Font(color=SUB, italic=True)
autosize(ws, COLS, row - 2)
ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}{max(row - 1, 1)}"

# ========================= Sheet 3: Not Tracked =============================
ws = wb.create_sheet("Not Tracked")
ws.sheet_view.showGridLines = False
COLS = [
    {"h": "Tracking status", "align": "center", "w": 15},
    {"h": "Severity", "align": "center", "w": 11},
    {"h": "Category", "wrap": True, "min": 16, "max": 22, "w": 20},
    {"h": "Finding", "wrap": True, "w": 34},
    {"h": "Rule ID", "min": 12, "max": 40},
    {"h": "Scanner", "w": 12},
    {"h": "Count", "align": "center", "w": 8},
    {"h": "Reason excluded", "wrap": True, "w": 52},
    {"h": "Why (context)", "wrap": True, "w": 52},
]
header(ws, COLS)
row = 2
for f in acceptable + false_pos:
    tri = f.get("triage") or {}
    label = "Acceptable" if status_of(f) == "acceptable" else "False positive"
    rf = PatternFill("solid", fgColor=NEUTRAL_SOFT)
    vals = {3: f.get("category", ""), 4: f.get("title", ""), 5: f.get("id", ""), 6: f.get("source", ""),
            7: f.get("count", 0), 8: tri.get("reason", ""), 9: f.get("why", "")}
    for c, v in vals.items():
        cell = ws.cell(row=row, column=c, value=v)
        cell.alignment = WRAP if COLS[c - 1].get("wrap") else TOP
        cell.border = BORDER; cell.fill = rf
    ac = ws.cell(row=row, column=1, value=label); ac.alignment = CTR; ac.border = BORDER
    ac.font = Font(bold=True, size=10, color="475569" if label == "Acceptable" else "845C05")
    ac.fill = PatternFill("solid", fgColor="E9EDF4" if label == "Acceptable" else "FCF1D0")
    chip(ws.cell(row=row, column=2), f["severity"].title(), *SEV_CHIP.get(f["severity"], ("EAEDF2", "475569")))
    row += 1
if row == 2:
    ws.cell(row=2, column=1, value="Nothing excluded — every finding is action-required.").font = Font(color=SUB, italic=True)
autosize(ws, COLS, row - 2)
ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}{max(row - 1, 1)}"

# ======================= Sheet 4: All Locations =============================
ws = wb.create_sheet("All Locations")
ws.sheet_view.showGridLines = False
COLS = [
    {"h": "Priority", "align": "center", "w": 9},
    {"h": "Severity", "align": "center", "w": 11},
    {"h": "Category", "wrap": True, "min": 16, "max": 22, "w": 20},
    {"h": "Finding", "wrap": True, "w": 40},
    {"h": "Rule ID", "min": 12, "max": 40},
    {"h": "File", "min": 20, "max": 50},
    {"h": "Line", "align": "center", "w": 8},
    {"h": "Link", "align": "center", "w": 10},
]
header(ws, COLS)
row = 2
for f in action:
    pr = priority_of(f)
    rf = PatternFill("solid", fgColor=PRI_SOFT.get(pr, NEUTRAL_SOFT))
    for loc in (f.get("locations") or []):
        vals = {3: f.get("category", ""), 4: f.get("title", ""), 5: f.get("id", ""),
                6: loc.get("path", ""), 7: loc.get("line", "")}
        for c, v in vals.items():
            cell = ws.cell(row=row, column=c, value=v)
            cell.alignment = WRAP if COLS[c - 1].get("wrap") else TOP
            cell.border = BORDER; cell.fill = rf
        lk = ws.cell(row=row, column=8); link(lk, "open ↗" if loc.get("url") else "", loc.get("url")); lk.alignment = CTR; lk.border = BORDER; lk.fill = rf
        solid_chip(ws.cell(row=row, column=1), pr, PRI_SOLID.get(pr, "5F6673"))
        chip(ws.cell(row=row, column=2), f["severity"].title(), *SEV_CHIP.get(f["severity"], ("EAEDF2", "475569")))
        row += 1
if row == 2:
    ws.cell(row=2, column=1, value="No action-required locations.").font = Font(color=SUB, italic=True)
autosize(ws, COLS, row - 2)
ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}{max(row - 1, 1)}"

wb.save(OUT)
print(f"wrote {OUT}: {len(action)} action-required, {len(acceptable)} acceptable, {len(false_pos)} false-positive.", file=sys.stderr)
